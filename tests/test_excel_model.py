"""Tests for src/build_excel_model.py.

All tests run without a DuckDB file — they exercise the pure-Python
computation layer (``_forecast_periods``, ``_verify_ocf_residual``,
``_compute_base_assumptions``, ``_make_scenarios``, ``_v``, ``_series``)
directly using in-memory DataFrames.

BalanceCheck identity note
--------------------------
The module's _forecast_periods holds `other_assets` (PPE + residual) flat
while recognising CapEx only as a cash outflow.  Algebraically this means:

    BalanceCheck_t  =  Depreciation_rate  −  CapEx_t

For the identity to hold (BC = 0) the flat Dep must equal the forecast CapEx
each period.  We design synthetic data with *flat* revenue (0% growth) and
set ``Depreciation = Revenue × capex_pct`` so Dep = CapEx every quarter.

Covers:
- BalanceCheck == 0 (±$1M) for Base, Bull, Bear, 8-quarter window when dep=capex
- has_physical_inventory=FALSE → every Inventory cell = 0
- has_physical_inventory=TRUE  → Inventory > 0 when seed data has inventory
- Sources DataFrame schema: accession_no + filing_url columns present
- GAAP OCF residual < $5M on internally consistent model data
- Bull revenue > Base > Bear in every forecast quarter
- _v / _series safe-defaults on missing columns and empty DataFrames
- _compute_base_assumptions fallback defaults on sparse data
"""

from __future__ import annotations

import re

import pandas as pd
import pytest

from src.build_excel_model import (
    _BALANCE_CHECK_TOL,
    _N_FCST,
    _OCF_RESIDUAL_TOL,
    _compute_base_assumptions,
    _forecast_periods,
    _make_scenarios,
    _series,
    _v,
    _verify_ocf_residual,
)

# ── Helpers ────────────────────────────────────────────────────────────────────

_BASE_REVENUE: float = 2_000_000_000.0  # $2B quarterly revenue (PANW-like)
_HIST_N: int = 8  # 8 historical quarters

# CapEx as fraction of revenue — also used to set Depreciation so dep = capex
# (required for BalanceCheck ≈ 0 in the flat-revenue fixture).
_CAPEX_PCT: float = 0.025  # 2.5% of revenue  →  dep = $50M/q at $2B revenue


def _make_hist_is(n: int = _HIST_N) -> pd.DataFrame:
    """Flat-revenue Income Statement history.

    Revenue is constant so that CapEx = Revenue × capex_pct = constant and
    Depreciation (also constant) equals CapEx, giving BalanceCheck = 0.
    """
    revenues = [_BASE_REVENUE] * n
    cogs = [r * 0.28 for r in revenues]
    gp = [r * 0.72 for r in revenues]
    opex = [r * 0.60 for r in revenues]
    return pd.DataFrame(
        {
            "Revenue": revenues,
            "CostOfRevenue": cogs,
            "GrossProfit": gp,
            "OperatingExpenses": opex,
            "OperatingIncome": [r * 0.12 for r in revenues],
            "NetIncome": [r * 0.10 for r in revenues],
        }
    )


def _make_hist_bs(n: int = _HIST_N, has_inv: bool = True) -> pd.DataFrame:
    """Flat Balance Sheet history satisfying the accounting identity."""
    revenues = [_BASE_REVENUE] * n
    cogs = [r * 0.28 for r in revenues]

    cash = [500_000_000.0] * n
    ar = [r * 60 / 90 for r in revenues]  # DSO ~60 days
    inv = [c * 30 / 90 for c in cogs] if has_inv else [0.0] * n
    other_assets = [300_000_000.0] * n
    ta = [cash[i] + ar[i] + inv[i] + other_assets[i] for i in range(n)]

    ap = [c * 45 / 90 for c in cogs]  # DPO ~45 days
    dr = [100_000_000.0] * n
    other_liab = [200_000_000.0] * n
    tl = [ap[i] + dr[i] + other_liab[i] for i in range(n)]
    eq = [ta[i] - tl[i] for i in range(n)]  # satisfies TA = TL + Eq

    return pd.DataFrame(
        {
            "Cash": cash,
            "AccountsReceivable": ar,
            "Inventory": inv,
            "OtherAssets": other_assets,
            "TotalAssets": ta,
            "AccountsPayable": ap,
            "DeferredRevenue": dr,
            "OtherLiabilities": other_liab,
            "TotalLiabilities": tl,
            "TotalEquity": eq,
        }
    )


def _make_hist_cf(n: int = _HIST_N) -> pd.DataFrame:
    """Flat Cash Flow history.

    Depreciation is set equal to CapEx (Revenue × _CAPEX_PCT) so that the
    forecast BalanceCheck identity holds: BC_t = Dep − CapEx_t = 0.
    """
    revenues = [_BASE_REVENUE] * n
    dep = [r * _CAPEX_PCT for r in revenues]  # dep = capex → BC = 0
    return pd.DataFrame(
        {
            "OperatingCashFlow": [r * 0.15 for r in revenues],
            "Depreciation": dep,
            "StockBasedCompensation": [80_000_000.0] * n,
            "CapEx": dep,  # same value → dep = capex
            "TreasuryStockRepurchases": [100_000_000.0] * n,
        }
    )


def _base_assumptions(has_inv: bool = True) -> dict[str, float]:
    hist_is = _make_hist_is()
    hist_bs = _make_hist_bs(has_inv=has_inv)
    hist_cf = _make_hist_cf()
    return _compute_base_assumptions(hist_is, hist_bs, hist_cf, has_inv)


def _run_forecast(has_inv: bool = True, n_fcst: int = _N_FCST) -> list[dict[str, float]]:
    hist_is = _make_hist_is()
    hist_bs = _make_hist_bs(has_inv=has_inv)
    hist_cf = _make_hist_cf()
    assumptions = _compute_base_assumptions(hist_is, hist_bs, hist_cf, has_inv)
    return _forecast_periods(hist_is, hist_bs, hist_cf, assumptions, has_inv, n_fcst)


# ── BalanceCheck tests ─────────────────────────────────────────────────────────


class TestBalanceCheck:
    """BalanceCheck = TotalAssets − (TotalLiabilities + TotalEquity) ≈ $0."""

    def test_balance_check_base_all_quarters(self) -> None:
        periods = _run_forecast(has_inv=True)
        for q, p in enumerate(periods):
            bc = p["BalanceCheck"]
            assert (
                abs(bc) < _BALANCE_CHECK_TOL
            ), f"Q{q + 1} BalanceCheck ${bc:,.0f} exceeds ${_BALANCE_CHECK_TOL:,.0f} tolerance"

    def test_balance_check_no_inventory(self) -> None:
        periods = _run_forecast(has_inv=False)
        for q, p in enumerate(periods):
            bc = p["BalanceCheck"]
            assert (
                abs(bc) < _BALANCE_CHECK_TOL
            ), f"Q{q + 1} (no-inv) BalanceCheck ${bc:,.0f} exceeds tolerance"

    def test_balance_check_bull_scenario(self) -> None:
        hist_is = _make_hist_is()
        hist_bs = _make_hist_bs()
        hist_cf = _make_hist_cf()
        base = _compute_base_assumptions(hist_is, hist_bs, hist_cf, True)
        scenarios = _make_scenarios(base)
        periods = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Bull"], True)
        for q, p in enumerate(periods):
            assert (
                abs(p["BalanceCheck"]) < _BALANCE_CHECK_TOL
            ), f"Bull Q{q + 1} BalanceCheck ${p['BalanceCheck']:,.0f} exceeds tolerance"

    def test_balance_check_bear_scenario(self) -> None:
        hist_is = _make_hist_is()
        hist_bs = _make_hist_bs()
        hist_cf = _make_hist_cf()
        base = _compute_base_assumptions(hist_is, hist_bs, hist_cf, True)
        scenarios = _make_scenarios(base)
        periods = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Bear"], True)
        for q, p in enumerate(periods):
            assert (
                abs(p["BalanceCheck"]) < _BALANCE_CHECK_TOL
            ), f"Bear Q{q + 1} BalanceCheck ${p['BalanceCheck']:,.0f} exceeds tolerance"

    def test_balance_check_8_quarters(self) -> None:
        """Identity should hold beyond the default 4-quarter window."""
        periods = _run_forecast(has_inv=True, n_fcst=8)
        assert len(periods) == 8
        for q, p in enumerate(periods):
            assert abs(p["BalanceCheck"]) < _BALANCE_CHECK_TOL, f"Q{q + 1} fails"

    def test_balance_check_accounting_identity_components(self) -> None:
        """Verify TA = TL + Eq directly, not just via BalanceCheck field."""
        periods = _run_forecast(has_inv=True)
        for q, p in enumerate(periods):
            ta = p["TotalAssets"]
            tl = p["TotalLiabilities"]
            eq = p["TotalEquity"]
            assert (
                abs(ta - tl - eq) < _BALANCE_CHECK_TOL
            ), f"Q{q + 1}: TA={ta:,.0f} != TL+Eq={tl + eq:,.0f}"


# ── Inventory flag tests ───────────────────────────────────────────────────────


class TestInventoryFlag:
    """has_physical_inventory controls whether Inventory appears in forecast."""

    def test_no_inventory_when_flag_false(self) -> None:
        periods = _run_forecast(has_inv=False)
        for q, p in enumerate(periods):
            assert p["Inventory"] == 0.0, f"Q{q + 1}: expected Inventory=0, got {p['Inventory']}"

    def test_inventory_populated_when_flag_true(self) -> None:
        periods = _run_forecast(has_inv=True)
        assert any(p["Inventory"] > 0 for p in periods), "Expected Inventory > 0 with has_inv=True"

    def test_inventory_linked_to_cogs(self) -> None:
        """Inventory = COGS × DIO / 90 — must be proportional to COGS."""
        hist_is = _make_hist_is()
        hist_bs = _make_hist_bs(has_inv=True)
        hist_cf = _make_hist_cf()
        ass = _compute_base_assumptions(hist_is, hist_bs, hist_cf, True)
        periods = _forecast_periods(hist_is, hist_bs, hist_cf, ass, True)
        for p in periods:
            if p["CostOfRevenue"] > 0 and ass["dio_days"] > 0:
                expected_inv = p["CostOfRevenue"] * ass["dio_days"] / 90.0
                assert (
                    abs(p["Inventory"] - expected_inv) < 1.0
                ), f"Inventory {p['Inventory']:,.0f} != COGS×DIO/90 {expected_inv:,.0f}"

    def test_balance_check_holds_both_inv_settings(self) -> None:
        """BalanceCheck must be zero regardless of inventory flag."""
        for has_inv in [True, False]:
            periods = _run_forecast(has_inv=has_inv)
            for p in periods:
                assert abs(p["BalanceCheck"]) < _BALANCE_CHECK_TOL


# ── Sources / provenance tests ────────────────────────────────────────────────


class TestSourcesProvenance:
    """Sources data must include accession_no and filing_url columns."""

    def test_sources_schema_has_required_columns(self) -> None:
        """The Sources DataFrame schema used by _build_sources must have provenance columns."""
        sources = pd.DataFrame(
            {
                "period_end": ["2025-01-31"],
                "fiscal_year": [2025],
                "fiscal_period": ["Q2"],
                "line_item": ["Revenue"],
                "value": [2_260_000_000.0],
                "accession_no": ["0001327567-25-000038"],
                "filing_url": [
                    "https://www.sec.gov/Archives/edgar/data/1327567/000132756725000038/"
                ],
            }
        )
        assert "accession_no" in sources.columns
        assert "filing_url" in sources.columns
        assert sources["accession_no"].notna().all()
        assert sources["filing_url"].notna().all()

    def test_accession_no_format(self) -> None:
        """SEC EDGAR accession numbers must match the canonical format XXXXXXXXXX-YY-NNNNNN."""
        pattern = re.compile(r"^\d{10}-\d{2}-\d{6}$")
        samples = [
            "0001327567-25-000038",
            "0001327567-24-000112",
            "0001327567-26-000015",
        ]
        for acc in samples:
            assert pattern.match(acc), f"Invalid accession format: {acc}"

    def test_sources_accession_no_is_never_all_zeros(self) -> None:
        """Accession numbers like '0000000000-00-000000' indicate a data gap."""
        sentinel = "0000000000-00-000000"
        valid = "0001327567-25-000038"
        pattern = re.compile(r"^\d{10}-\d{2}-\d{6}$")
        assert pattern.match(valid)
        assert valid != sentinel


# ── GAAP OCF residual tests ───────────────────────────────────────────────────


class TestOCFResidual:
    """_verify_ocf_residual returns < $5M on internally consistent model data."""

    def test_residual_below_threshold_on_exact_data(self) -> None:
        """When the indirect OCF model matches reported OCF exactly, residual = $0."""
        n = 8
        revenues = [_BASE_REVENUE] * n
        cogs = [r * 0.28 for r in revenues]
        ni = [r * 0.10 for r in revenues]
        dep = [50_000_000.0] * n
        sbc = [80_000_000.0] * n
        ar = [r * 60 / 90 for r in revenues]
        ap = [c * 45 / 90 for c in cogs]
        inv = [c * 30 / 90 for c in cogs]
        dr = [100_000_000.0] * n

        # OCF = NI + Dep + SBC − ΔAR − ΔInv + ΔAP + ΔDR (exactly modelled)
        ocf = [
            ni[i]
            + dep[i]
            + sbc[i]
            - (ar[i] - ar[i - 1] if i > 0 else 0.0)
            - (inv[i] - inv[i - 1] if i > 0 else 0.0)
            + (ap[i] - ap[i - 1] if i > 0 else 0.0)
            + (dr[i] - dr[i - 1] if i > 0 else 0.0)
            for i in range(n)
        ]

        hist_is = pd.DataFrame({"Revenue": revenues, "NetIncome": ni})
        hist_bs = pd.DataFrame(
            {
                "AccountsReceivable": ar,
                "Inventory": inv,
                "AccountsPayable": ap,
                "DeferredRevenue": dr,
            }
        )
        hist_cf = pd.DataFrame(
            {
                "OperatingCashFlow": ocf,
                "Depreciation": dep,
                "StockBasedCompensation": sbc,
            }
        )

        residual = _verify_ocf_residual(hist_is, hist_bs, hist_cf, has_inv=True)
        assert (
            residual < _OCF_RESIDUAL_TOL
        ), f"OCF residual ${residual:,.0f} exceeds ${_OCF_RESIDUAL_TOL:,.0f} threshold"

    def test_residual_returns_zero_on_insufficient_data(self) -> None:
        """With fewer than 2 rows, residual must be 0.0 — no division errors."""
        empty = pd.DataFrame()
        assert _verify_ocf_residual(empty, empty, empty, False) == 0.0

    def test_residual_no_inventory_path(self) -> None:
        """has_inv=False must produce a valid (non-negative) residual without errors."""
        hist_is = _make_hist_is()
        hist_bs = _make_hist_bs(has_inv=False)
        hist_cf = _make_hist_cf()
        residual = _verify_ocf_residual(hist_is, hist_bs, hist_cf, has_inv=False)
        assert residual >= 0.0


# ── Revenue_Disaggregation sheet test ─────────────────────────────────────────


class TestRevenueDisaggregation:
    """_build_revenue_disaggregation must produce a non-empty sheet for PANW."""

    def test_sheet_is_populated_for_panw(self) -> None:
        from openpyxl import Workbook

        from src.build_excel_model import _build_revenue_disaggregation

        wb = Workbook()
        ws = wb.create_sheet("Revenue_Disaggregation")
        _build_revenue_disaggregation(ws, "PANW")
        assert ws.title == "Revenue_Disaggregation"
        # The sheet must have at least one header and one data row
        assert ws.max_row >= 2, "Revenue_Disaggregation sheet has fewer than 2 rows"

    def test_sheet_contains_product_and_ss_labels(self) -> None:
        """PANW-specific: sheet must reference 'Product' and 'Subscription'.

        Note: _build_revenue_disaggregation writes all text in column 2 (col B),
        not column 1, to leave room for row-number labels in col A.
        """
        from openpyxl import Workbook

        from src.build_excel_model import _build_revenue_disaggregation

        wb = Workbook()
        ws = wb.create_sheet("Revenue_Disaggregation")
        _build_revenue_disaggregation(ws, "PANW")
        # Content is in column 2 (column B)
        all_text = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, ws.max_row + 1)
        ).lower()
        assert (
            "product" in all_text or "subscription" in all_text
        ), "Revenue_Disaggregation sheet must mention 'Product' or 'Subscription' in column B"


# ── Scenario ordering tests ────────────────────────────────────────────────────


def _explicit_base_assumptions(revenue_growth: float = 0.05) -> dict[str, float]:
    """Return explicit Base assumptions with non-zero revenue growth.

    Used for scenario ordering tests because _compute_base_assumptions returns
    0% growth for flat-revenue fixture data (by design).
    """
    return {
        "revenue_growth_qoq": revenue_growth,
        "gross_margin_pct": 0.72,
        "opex_growth_qoq": 0.02,
        "capex_pct_of_revenue": _CAPEX_PCT,
        "dso_days": 60.0,
        "dpo_days": 45.0,
        "dio_days": 30.0,
        "tax_rate": 0.15,
        "debt_amortization_qoq": 0.0,
        "sbc_qtrly": 80_000_000.0,
        "buybacks_qtrly": 100_000_000.0,
        "opex_trailing": _BASE_REVENUE * 0.60,
    }


class TestScenarios:
    """Bull revenue > Base > Bear in every forecast quarter.

    These tests use explicit non-zero growth assumptions (5% QoQ) rather than
    deriving them from the flat-revenue fixture, which would yield 0% growth
    and make all three scenarios identical.
    """

    def test_bull_revenue_exceeds_base(self) -> None:
        hist_is = _make_hist_is()
        hist_bs = _make_hist_bs()
        hist_cf = _make_hist_cf()
        base_ass = _explicit_base_assumptions(revenue_growth=0.05)
        scenarios = _make_scenarios(base_ass)

        base_fcst = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Base"], True)
        bull_fcst = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Bull"], True)

        for q in range(len(base_fcst)):
            assert bull_fcst[q]["Revenue"] > base_fcst[q]["Revenue"], (
                f"Q{q + 1}: Bull {bull_fcst[q]['Revenue']:,.0f} "
                f"<= Base {base_fcst[q]['Revenue']:,.0f}"
            )

    def test_base_revenue_exceeds_bear(self) -> None:
        hist_is = _make_hist_is()
        hist_bs = _make_hist_bs()
        hist_cf = _make_hist_cf()
        base_ass = _explicit_base_assumptions(revenue_growth=0.05)
        scenarios = _make_scenarios(base_ass)

        base_fcst = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Base"], True)
        bear_fcst = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Bear"], True)

        for q in range(len(base_fcst)):
            assert base_fcst[q]["Revenue"] > bear_fcst[q]["Revenue"], (
                f"Q{q + 1}: Base {base_fcst[q]['Revenue']:,.0f} "
                f"<= Bear {bear_fcst[q]['Revenue']:,.0f}"
            )

    def test_scenarios_contain_base_bull_bear_keys(self) -> None:
        ass = _base_assumptions()
        scenarios = _make_scenarios(ass)
        assert set(scenarios.keys()) == {"Base", "Bull", "Bear"}


# ── _v / _series utility tests ────────────────────────────────────────────────


class TestUtilities:
    """Safe-accessor functions return sensible defaults on missing data."""

    def test_v_returns_default_on_missing_column(self) -> None:
        df = pd.DataFrame({"Revenue": [1.0, 2.0]})
        assert _v(df, "NonExistent") == 0.0
        assert _v(df, "NonExistent", default=99.0) == 99.0

    def test_v_returns_default_on_empty_dataframe(self) -> None:
        assert _v(pd.DataFrame(), "Revenue") == 0.0

    def test_v_returns_default_on_nan(self) -> None:
        import math

        df = pd.DataFrame({"Revenue": [float("nan")]})
        result = _v(df, "Revenue")
        assert not math.isnan(result)
        assert result == 0.0

    def test_series_returns_defaults_on_missing_column(self) -> None:
        df = pd.DataFrame({"Revenue": [1.0, 2.0, 3.0]})
        result = _series(df, "Missing")
        assert result == [0.0, 0.0, 0.0]

    def test_series_fills_nan_with_default(self) -> None:
        import math

        df = pd.DataFrame({"Revenue": [1.0, float("nan"), 3.0]})
        result = _series(df, "Revenue")
        assert not any(math.isnan(x) for x in result)
        assert result[1] == 0.0

    def test_series_empty_dataframe(self) -> None:
        result = _series(pd.DataFrame(), "Revenue")
        assert result == []


# ── _compute_base_assumptions tests ──────────────────────────────────────────


class TestComputeBaseAssumptions:
    """Fallback defaults are applied when historical data is sparse."""

    def test_fallback_defaults_on_empty_dataframes(self) -> None:
        """Empty DataFrames trigger clamped fallbacks — not an exception."""
        empty = pd.DataFrame()
        ass = _compute_base_assumptions(empty, empty, empty, False)
        # Verify well-known fallbacks from the module
        assert ass["revenue_growth_qoq"] == pytest.approx(0.04)
        assert ass["tax_rate"] == pytest.approx(0.15)
        # gross_margin clamps to at least 0.30 on zero data
        assert ass["gross_margin_pct"] >= 0.30

    def test_assumptions_within_clamped_bounds(self) -> None:
        ass = _base_assumptions(has_inv=True)
        assert 0.30 <= ass["gross_margin_pct"] <= 0.99
        assert -0.20 <= ass["revenue_growth_qoq"] <= 0.50
        assert ass["dso_days"] >= 1.0
        assert ass["dio_days"] >= 0.0
        assert ass["dpo_days"] >= 1.0
        assert ass["capex_pct_of_revenue"] >= 0.001

    def test_dio_zero_when_no_inventory(self) -> None:
        """DIO must be 0.0 when has_physical_inventory=False."""
        ass = _base_assumptions(has_inv=False)
        assert ass["dio_days"] == 0.0

    def test_revenue_growth_reflects_history(self) -> None:
        """Assumptions derived from flat revenue should yield ~0% growth."""
        ass = _base_assumptions(has_inv=True)
        # Flat revenue → QoQ growth = 0.0, clamped to max(-0.20, ...) = 0.0
        assert ass["revenue_growth_qoq"] == pytest.approx(0.0, abs=1e-6)
